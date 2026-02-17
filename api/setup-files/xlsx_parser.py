# api/services/documents/parsers/xlsx_parser.py
"""
Parser for Microsoft Excel spreadsheets (.xlsx, .xls, .csv).
"""

import csv
import logging
import subprocess
import tempfile
from io import StringIO
from pathlib import Path
from typing import BinaryIO, Optional, Union
import time

from ..base import (
    BaseParser, DocumentContent, DocumentMetadata,
    FileType, ParserResult, TextChunk
)

logger = logging.getLogger(__name__)


class XlsxParser(BaseParser):
    """Parser for Excel spreadsheets and CSV files."""
    
    supported_types = [FileType.XLSX, FileType.XLS, FileType.CSV]
    
    async def parse(
        self,
        file_path: Union[str, Path],
        chunk_size: Optional[int] = None,
        chunk_overlap: Optional[int] = None,
    ) -> ParserResult:
        """Parse a spreadsheet from file path."""
        start_time = time.time()
        file_path = Path(file_path)
        
        try:
            with open(file_path, "rb") as f:
                file_data = f.read()
            
            result = await self._parse_content(
                file_data,
                file_path.name,
                chunk_size,
                chunk_overlap,
            )
            result.processing_time_ms = (time.time() - start_time) * 1000
            return result
            
        except Exception as e:
            logger.error(f"Error parsing spreadsheet {file_path}: {e}")
            return ParserResult(
                success=False,
                error=str(e),
                processing_time_ms=(time.time() - start_time) * 1000,
            )
    
    async def parse_bytes(
        self,
        file_data: BinaryIO,
        filename: str,
        chunk_size: Optional[int] = None,
        chunk_overlap: Optional[int] = None,
    ) -> ParserResult:
        """Parse a spreadsheet from bytes."""
        start_time = time.time()
        
        try:
            data = file_data.read()
            result = await self._parse_content(
                data,
                filename,
                chunk_size,
                chunk_overlap,
            )
            result.processing_time_ms = (time.time() - start_time) * 1000
            return result
            
        except Exception as e:
            logger.error(f"Error parsing spreadsheet bytes: {e}")
            return ParserResult(
                success=False,
                error=str(e),
                processing_time_ms=(time.time() - start_time) * 1000,
            )
    
    async def _parse_content(
        self,
        file_data: bytes,
        filename: str,
        chunk_size: Optional[int],
        chunk_overlap: Optional[int],
    ) -> ParserResult:
        """Parse spreadsheet content."""
        file_type = self.detect_file_type(filename)
        file_hash = self.compute_file_hash(file_data)
        warnings = []
        
        # CSV can be parsed directly
        if file_type == FileType.CSV:
            return await self._parse_csv(
                file_data, filename, file_hash, chunk_size, chunk_overlap
            )
        
        # Excel files need temp file
        with tempfile.NamedTemporaryFile(
            suffix=f".{file_type.value}", delete=False
        ) as tmp:
            tmp.write(file_data)
            tmp_path = Path(tmp.name)
        
        try:
            # Convert .xls to .xlsx if needed
            if file_type == FileType.XLS:
                converted_path = await self._convert_xls_to_xlsx(tmp_path)
                if converted_path:
                    tmp_path = converted_path
                else:
                    warnings.append("Could not convert .xls to .xlsx")
            
            # Extract content
            sheets_text, chunks, tables = await self._extract_with_openpyxl(
                tmp_path, chunk_size, chunk_overlap
            )
            
            if not sheets_text:
                # Fallback to pandas
                sheets_text, chunks, tables = await self._extract_with_pandas(
                    tmp_path, chunk_size, chunk_overlap
                )
            
            # Extract metadata
            metadata = await self._extract_metadata(
                tmp_path, file_data, filename, file_type, file_hash
            )
            
            content = DocumentContent(
                full_text=sheets_text,
                chunks=chunks,
                tables=tables,
            )
            
            return ParserResult(
                success=True,
                metadata=metadata,
                content=content,
                warnings=warnings,
            )
            
        finally:
            tmp_path.unlink(missing_ok=True)
            if file_type == FileType.XLS:
                xlsx_path = tmp_path.with_suffix(".xlsx")
                xlsx_path.unlink(missing_ok=True)
    
    async def _parse_csv(
        self,
        file_data: bytes,
        filename: str,
        file_hash: str,
        chunk_size: Optional[int],
        chunk_overlap: Optional[int],
    ) -> ParserResult:
        """Parse CSV file."""
        try:
            # Try to decode with different encodings
            text = None
            for encoding in ["utf-8", "utf-8-sig", "latin-1", "cp1252"]:
                try:
                    text = file_data.decode(encoding)
                    break
                except UnicodeDecodeError:
                    continue
            
            if text is None:
                return ParserResult(
                    success=False,
                    error="Could not decode CSV file",
                )
            
            # Parse CSV
            reader = csv.reader(StringIO(text))
            rows = list(reader)
            
            if not rows:
                return ParserResult(
                    success=False,
                    error="Empty CSV file",
                )
            
            # Convert to text representation
            text_rows = []
            header = rows[0] if rows else []
            
            for i, row in enumerate(rows):
                if i == 0:
                    text_rows.append("Headers: " + " | ".join(row))
                else:
                    # Create key-value representation for better context
                    row_text = []
                    for j, cell in enumerate(row):
                        if cell.strip():
                            col_name = header[j] if j < len(header) else f"col_{j}"
                            row_text.append(f"{col_name}: {cell}")
                    if row_text:
                        text_rows.append(f"Row {i}: " + ", ".join(row_text))
            
            full_text = "\n".join(text_rows)
            chunks = self.chunk_text(full_text, "csv_chunk", chunk_size, chunk_overlap)
            
            metadata = DocumentMetadata(
                filename=filename,
                file_type=FileType.CSV,
                file_size=len(file_data),
                file_hash=file_hash,
                custom={"row_count": len(rows), "column_count": len(header)},
            )
            
            content = DocumentContent(
                full_text=full_text,
                chunks=chunks,
                tables=[{"sheet": "data", "rows": rows}],
            )
            
            return ParserResult(
                success=True,
                metadata=metadata,
                content=content,
            )
            
        except Exception as e:
            logger.error(f"Error parsing CSV: {e}")
            return ParserResult(success=False, error=str(e))
    
    async def _convert_xls_to_xlsx(self, xls_path: Path) -> Optional[Path]:
        """Convert .xls to .xlsx using LibreOffice."""
        try:
            output_dir = xls_path.parent
            result = subprocess.run(
                [
                    "soffice", "--headless", "--convert-to", "xlsx",
                    "--outdir", str(output_dir), str(xls_path)
                ],
                capture_output=True,
                timeout=120,
            )
            if result.returncode == 0:
                xlsx_path = xls_path.with_suffix(".xlsx")
                if xlsx_path.exists():
                    return xlsx_path
        except Exception as e:
            logger.warning(f"Failed to convert .xls to .xlsx: {e}")
        return None
    
    async def _extract_with_openpyxl(
        self,
        file_path: Path,
        chunk_size: Optional[int],
        chunk_overlap: Optional[int],
    ) -> tuple[str, list[TextChunk], list[dict]]:
        """Extract content using openpyxl."""
        try:
            from openpyxl import load_workbook
        except ImportError:
            logger.warning("openpyxl not installed")
            return "", [], []
        
        try:
            wb = load_workbook(str(file_path), data_only=True, read_only=True)
            
            all_text = []
            chunks = []
            tables = []
            chunk_idx = 0
            
            for sheet_name in wb.sheetnames:
                sheet = wb[sheet_name]
                sheet_rows = []
                
                for row in sheet.iter_rows(values_only=True):
                    # Filter out completely empty rows
                    if any(cell is not None for cell in row):
                        row_values = [str(cell) if cell is not None else "" for cell in row]
                        sheet_rows.append(row_values)
                
                if sheet_rows:
                    # Create text representation
                    header = sheet_rows[0] if sheet_rows else []
                    sheet_text_parts = [f"[Sheet: {sheet_name}]"]
                    
                    for i, row in enumerate(sheet_rows):
                        if i == 0:
                            sheet_text_parts.append("Headers: " + " | ".join(row))
                        else:
                            row_text = []
                            for j, cell in enumerate(row):
                                if cell.strip():
                                    col_name = header[j] if j < len(header) else f"col_{j}"
                                    row_text.append(f"{col_name}: {cell}")
                            if row_text:
                                sheet_text_parts.append(", ".join(row_text))
                    
                    sheet_text = "\n".join(sheet_text_parts)
                    all_text.append(sheet_text)
                    
                    # Create chunk for this sheet
                    chunks.append(TextChunk(
                        content=sheet_text,
                        source=f"sheet_{sheet_name}",
                        chunk_index=chunk_idx,
                    ))
                    chunk_idx += 1
                    
                    tables.append({
                        "sheet": sheet_name,
                        "rows": sheet_rows,
                    })
            
            wb.close()
            
            full_text = "\n\n".join(all_text)
            
            # Re-chunk if needed
            if chunk_size:
                chunks = self.chunk_text(full_text, "chunk", chunk_size, chunk_overlap)
            
            return full_text, chunks, tables
            
        except Exception as e:
            logger.error(f"Error with openpyxl: {e}")
            return "", [], []
    
    async def _extract_with_pandas(
        self,
        file_path: Path,
        chunk_size: Optional[int],
        chunk_overlap: Optional[int],
    ) -> tuple[str, list[TextChunk], list[dict]]:
        """Extract content using pandas as fallback."""
        try:
            import pandas as pd
        except ImportError:
            logger.warning("pandas not installed")
            return "", [], []
        
        try:
            sheets = pd.read_excel(str(file_path), sheet_name=None)
            
            all_text = []
            chunks = []
            tables = []
            chunk_idx = 0
            
            for sheet_name, df in sheets.items():
                if df.empty:
                    continue
                
                sheet_text = f"[Sheet: {sheet_name}]\n{df.to_string()}"
                all_text.append(sheet_text)
                
                chunks.append(TextChunk(
                    content=sheet_text,
                    source=f"sheet_{sheet_name}",
                    chunk_index=chunk_idx,
                ))
                chunk_idx += 1
                
                tables.append({
                    "sheet": sheet_name,
                    "rows": [df.columns.tolist()] + df.values.tolist(),
                })
            
            full_text = "\n\n".join(all_text)
            
            if chunk_size:
                chunks = self.chunk_text(full_text, "chunk", chunk_size, chunk_overlap)
            
            return full_text, chunks, tables
            
        except Exception as e:
            logger.error(f"Error with pandas: {e}")
            return "", [], []
    
    async def _extract_metadata(
        self,
        file_path: Path,
        file_data: bytes,
        filename: str,
        file_type: FileType,
        file_hash: str,
    ) -> DocumentMetadata:
        """Extract spreadsheet metadata."""
        metadata = DocumentMetadata(
            filename=filename,
            file_type=file_type,
            file_size=len(file_data),
            file_hash=file_hash,
        )
        
        try:
            from openpyxl import load_workbook
            wb = load_workbook(str(file_path), read_only=True)
            
            metadata.sheet_count = len(wb.sheetnames)
            
            # Document properties
            props = wb.properties
            if props:
                metadata.title = props.title
                metadata.author = props.creator
                metadata.subject = props.subject
                metadata.created_at = props.created
                metadata.modified_at = props.modified
            
            wb.close()
            
        except Exception as e:
            logger.warning(f"Could not extract Excel metadata: {e}")
        
        return metadata
